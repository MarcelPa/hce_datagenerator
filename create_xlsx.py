import argparse
import copy
import csv
import datetime
import random
import string
from typing import Any, List, Dict

from faker import Faker
from faker_bloodpanel import BloodPanelProvider
import openpyxl
from openpyxl.styles import PatternFill

def create_patients(fake: Faker, n: int) -> List[Dict[str, Any]]:
    ''' Generate a list of n patients. '''
    # get today to calculate visit dates
    today = datetime.date.today()

    patients = []
    for i in range(n):
        profile = fake.simple_profile()
        address = profile['address'].split('\n')
        patients.append({
            'naam': profile['name'],
            'patient_id': f'{(i+1) * 41232 % 100003:05d}', # a seemingly random, non consecutive patient id
            'datum': (today - datetime.timedelta(days=random.randint(1, 160))).isoformat(),
            'geboortedatum': profile['birthdate'].isoformat(),
            'straat': address[0],
            'stad': ' '.join(address[1:]),
        })
    return patients

def generate_record(samples: Dict[str, Any], fake: Faker) -> Dict[str, Any]:
    ''' Generate a record containing a random sample using Faker. '''

    # start and populate a new record
    record = {}
    sample_type = random.choice(list(samples.keys()))
    record['type'] = sample_type
    record['genomen'], record['ingang'] = sorted((fake.time(), fake.time()))
    record = record | samples[sample_type]()

    # instead of True / False, we would like to have 'positief' / 'negatief'
    if sample_type == 'uitstrijkje':
        record['Covid-PCR'] = 'positief' if record['Covid-PCR'] else 'negatief'

    return record

def to_excel(sheet: List[Dict[str, Any]], keys: List[str], column_widths: List[int]) -> openpyxl.Workbook:
    ''' Transform a list of records to an openpyxl.Workbook. '''
    # create a new excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Labor'

    for pos, key in enumerate(keys):
        ws[f'{string.ascii_uppercase[pos]}1'] = key.capitalize()
        ws.column_dimensions[f'{string.ascii_uppercase[pos]}'].width = column_widths[pos]

    # add the data rows
    for row, record in enumerate(small_sheet):
        for key in record.keys():
            pos = keys.index(key)
            ws[f'{string.ascii_uppercase[pos]}{row+2}'] = record[key]

    return wb

if __name__ == '__main__':
    # parse the command line arguments
    parser = argparse.ArgumentParser(description = 'Generate fake data for the Covid-19 laboratory')
    parser.add_argument('-s', '--small', help = 'generate a small sheet', type = int, default = 10)
    parser.add_argument('-l', '--large', help = 'generate a large sheet', type = int, default = 5000)
    parser.add_argument('-o', '--output', help = 'output file name, will be prepended with small_ and large_', type = str, required = True)
    args = parser.parse_args()

    # check arguments
    if args.small > args.large:
        raise ValueError('small sheet size cannot be larger than large sheet size')
    if args.small < 0 or args.large < 0:
        raise ValueError('sheet sizes cannot be negative')
    if args.large == 0:
        raise ValueError('large sheet size cannot be zero')
    if args.output == '':
        raise ValueError('output file name cannot be empty')
    if args.output.endswith('.xlsx'):
        args.output = args.output[:-5]

    # set the column names and widths
    keys = ['naam', 'patient_id', 'datum', 'geboortedatum', 'straat', 'stad', 'type', 'genomen', 'ingang', 'Sodium', 'Potassium', 'Chloride', 'Bicarbonate', 'Urea', 'Magnesium', 'Total calcium', 'Hemoglobin', 'Covid-PCR']
    column_widths = [18] + [14] * 3 + [18] * 2 + [14] * (len(keys) - 6) # header widths: name + others + street, city + rest

    # map the samples to a generator function
    samples = {
        'bloed': lambda: fake.simple_blood_panel(add_units = True),
        'uitstrijkje': lambda: fake.covid_test(),
    }

    # create a fake data generator
    fake = Faker()
    fake = Faker('nl_BE')
    fake.add_provider(BloodPanelProvider)

    # create a list of patients
    patients = create_patients(fake, int(args.large * .8))

    # check if a small file shall be created
    small_sheet = []
    if args.small > 0:
        # generate the small sheet: pick a few patient at random and generate samples for them
        small_i = int(args.small * .8) # 80% unique patients, 20% duplicates
        for patient in patients[:small_i] + random.sample(patients[:small_i], k = int(args.small * .2)):
            record = copy.deepcopy(patient)
            record = record | generate_record(samples, fake)
            small_sheet.append(record)

        # transform to excel and save it to a file
        wb = to_excel(small_sheet, keys, column_widths)
        wb.save(f'{datetime.date.today().isoformat()}_{args.output}_small.xlsx')
        print(f'Created {datetime.date.today().isoformat()}_{args.output}_small.xlsx')

    # generate the large sheet: pick all patients and generate samples for them
    large_sheet = small_sheet # we are just adding new records to the small sheet
    for patient in patients + random.sample(patients, k = int(args.large * .2) - args.small):
        record = copy.deepcopy(patient)
        record = record | generate_record(samples, fake)
        large_sheet.append(record)

    # introduce some noise to the large sheet
    for i in random.choices(list(range(args.small, args.large)), k = int(args.large * .05)):
        # define a list of 'noisable' keys
        noise_key = random.choice(['geboortedatum', 'type', 'Covid-PCR'])
        noise_map = {
            'geboortedatum': lambda cur: (datetime.date.fromisoformat(cur) - datetime.timedelta(days=random.randint(365 * 75, 365 * 150))).isoformat(),
            'type': lambda cur: 'bloed' if cur == 'uitstrijkje' else 'uitstrijkje',
            'Covid-PCR': lambda cur: random.choice(['pos', 'post', 'y', 'prositif', 'neg', '-', 'n', 'non'])
        }
        large_sheet[i][noise_key] = noise_map[noise_key](large_sheet[i][noise_key] if noise_key in large_sheet[i].keys() else None)

    # transform to excel and save it to a file
    wb = to_excel(large_sheet, keys, column_widths)
    wb.save(f'{datetime.date.today().isoformat()}_{args.output}_large.xlsx')
    print(f'Created {datetime.date.today().isoformat()}_{args.output}_large.xlsx')
    exit(0)
